import openpyxl
import os

class ExcelDataHandler:
    """A Python library to handle Excel data access using openpyxl."""
    
    def __init__(self):
        self.workbooks = {}

    def get_test_data_cell(self, file_path, sheet_name, test_case_no, column_name, start_row):
        """Reads a specific cell value based on a TestCase name and Column header."""
        
        # 1. Open/Load Workbook (Cached or New)
        if file_path not in self.workbooks:
            if not os.path.exists(file_path):
                 raise FileNotFoundError(f"Excel file not found at: {file_path}")
            self.workbooks[file_path] = openpyxl.load_workbook(file_path, data_only=True)
            
        wb = self.workbooks[file_path]
        
        # 2. Select Worksheet
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Worksheet '{sheet_name}' does not exist in '{file_path}'. Available sheets: {', '.join(wb.sheetnames)}")
            
        sheet = wb[sheet_name]
        
        # 3. Find Column Index (Header Row is start row)
        header_row_excel_index = int(start_row)
        col_index_excel = -1
        
        # Convert search string to uppercase and strip for robust comparison
        expected_column_name_search = column_name.strip().upper() 
        
        found_headers = [] # <--- NEW: List to collect all headers found
        
        # Iterate over cells in the header row
        for col in range(1, sheet.max_column + 1):
            header_cell_value = sheet.cell(row=header_row_excel_index, column=col).value
            
            # Store the found header (stripped) for the error message
            stripped_header = str(header_cell_value).strip() if header_cell_value is not None else ""
            found_headers.append(stripped_header)

            # Check for a match using case-insensitive comparison
            if stripped_header.upper() == expected_column_name_search:
                col_index_excel = col
                break
        
        if col_index_excel == -1:
            # <--- CRITICAL: Modified Error Message to show all headers found!
            raise ValueError(f"Column '{column_name}' not found in sheet '{sheet_name}'. "
                             f"The code searched row {header_row_excel_index}. "
                             f"Headers found in this row were: {found_headers}")

        # 4. Find Test Case Row and Retrieve Data
        
        for row_index_excel in range(header_row_excel_index + 1, sheet.max_row + 1):
            tc_name_cell_value = sheet.cell(row=row_index_excel, column=1).value
            
            if str(tc_name_cell_value).strip() == test_case_no.strip():
                data_cell = sheet.cell(row=row_index_excel, column=col_index_excel).value
                return str(data_cell) if data_cell is not None else ""
        
        raise LookupError(f"Test Case ID '{test_case_no}' not found in Column 1 of sheet '{sheet_name}'.")

    def close_all_workbooks(self):
        """Clean up by closing all open workbooks."""
        for wb in self.workbooks.values():
            try:
                wb.close()
            except Exception:
                pass 
        self.workbooks = {}
